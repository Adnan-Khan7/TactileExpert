#!/usr/bin/env python3
"""
TactileEval — Multi-Model Scoring Pipeline  (Appendix C implementation)

Runs all evaluator models over a dataset split, computes model consensus vs.
turker labels, flags disagreements, and outputs a review spreadsheet.

Usage
-----
  # GPU node (full 4-model scoring including VLM):
  python code/review/score_pipeline.py --split test

  # Single-model scoring (consensus = that model's vote alone):
  python code/review/score_pipeline.py --split train --models clip

  # Specific split and custom output directory:
  python code/review/score_pipeline.py --split val --out-dir review/val

Output
------
  review/{split}/scored_{split}.xlsx   — spreadsheet for manual review
  review/{split}/scoring_summary.txt   — conflict counts per option

Spreadsheet columns
-------------------
  pair_id, option, task_family, turker_vf, current_label,
  <model>_score for each model passed via --models,
  model_consensus, conflict_flag, default_label, user_label,
  nat_path, tac_path
"""

import argparse
import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill, Font
from tqdm import tqdm

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
HERE         = Path(__file__).resolve().parent
CODE_DIR     = HERE.parent
PROJECT_ROOT = CODE_DIR.parent

sys.path.insert(0, str(CODE_DIR))

from models.constants import ALL_OPTIONS
from models.evaluators import CLIPProbeEvaluator, TwoStreamEvaluator, VLMEvaluator

SPLITS_DIR  = PROJECT_ROOT / "TactileEval_Context_And_Data" / "data" / "splits"
IMAGES_DIR  = PROJECT_ROOT / "images"
MODELS_DIR  = PROJECT_ROOT / "TactileEval_Context_And_Data" / "models"
VLM_CKPT    = PROJECT_ROOT / "output" / "best_checkpoint"
REVIEW_DIR  = PROJECT_ROOT / "review"

# ---------------------------------------------------------------------------
# Label logic (Appendix C)
# ---------------------------------------------------------------------------
_CONSENSUS_THRESHOLD = 0.5  # probability boundary for each model's vote

def _model_consensus(scores: dict[str, float | None]) -> str:
    """
    Given model scores for one option (None if model not run), return consensus.
    Majority rule: ≥ ceil(n/2 + 1) models must agree for a definite consensus.
    With 4 models → 3 needed; with 3 models → 2 needed.
    Returns NO_MODEL when no model supports the option.
    """
    available = [(name, s) for name, s in scores.items() if s is not None]
    if not available:
        return "NO_MODEL"
    n       = len(available)
    n_pos   = sum(1 for _, s in available if s > _CONSENSUS_THRESHOLD)
    n_neg   = n - n_pos
    needed  = (n // 2) + 1
    if n_pos >= needed:
        return "1"
    if n_neg >= needed:
        return "0"
    return "SPLIT"


def _turker_label(vf: float | None) -> str:
    if vf is None:
        return "UNKNOWN"
    if vf >= 0.85:
        return "1"
    if vf <= 0.15:
        return "0"
    return "BORDERLINE"


def _conflict_flag(turker: str, consensus: str) -> bool:
    if consensus == "NO_MODEL":
        return False   # no model covers this option — can't assess
    if consensus == "SPLIT":
        return True
    if turker in ("1", "0") and turker != consensus:
        return True
    return False


def _conflict_flag_vf(vf: float | None) -> bool:
    """Flag rows where some turkers saw a defect but majority did not (0 < vf < 0.5).
    These are potential missed positives in the training labels."""
    return vf is not None and 0.0 < vf < 0.5


def _default_label(turker: str, current_label: int, consensus: str) -> str:
    if consensus == "NO_MODEL":
        return str(current_label)            # keep current label, no model to challenge it
    if turker in ("1", "0"):
        return turker                        # trust high-confidence turker label
    if consensus in ("1", "0"):
        return f"{consensus} (tentative)"    # model majority, no strong turker signal
    return ""                                # needs human review


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def load_split(split: str) -> list[dict]:
    path = SPLITS_DIR / f"{split}.jsonl"
    if not path.exists():
        raise FileNotFoundError(f"Split not found: {path}")
    with open(path) as f:
        return [json.loads(l) for l in f if l.strip()]


# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------
def score_split(
    records: list[dict],
    clip:    CLIPProbeEvaluator | None,
    resnet:  TwoStreamEvaluator | None,
    vit:     TwoStreamEvaluator | None,
    vlm:     VLMEvaluator | None,
) -> dict[tuple[str, str], dict[str, dict[str, float | None]]]:
    """
    Returns pair_scores[(nat_rel, tac_rel)][option][model_name] = prob_yes.
    Any of the four model args may be None to skip that model entirely.
    """
    unique_pairs = list({(r["natural_image"], r["tactile_image"]) for r in records})
    pair_scores: dict[tuple, dict] = {}

    print(f"\nScoring {len(unique_pairs)} unique image pairs...")

    empty = {opt: None for opt in ALL_OPTIONS}
    for nat_rel, tac_rel in tqdm(unique_pairs, desc="pairs"):
        nat_abs = str(IMAGES_DIR / nat_rel)
        tac_abs = str(IMAGES_DIR / tac_rel)

        try:
            clip_probs   = clip.score_pair(nat_abs, tac_abs)   if clip   else empty
            resnet_probs = resnet.score_pair(nat_abs, tac_abs) if resnet else empty
            vit_probs    = vit.score_pair(nat_abs, tac_abs)    if vit    else empty
            vlm_probs    = vlm.score_pair(nat_abs, tac_abs)    if vlm    else None
        except Exception as e:
            print(f"  [warn] error scoring {nat_rel}: {e}")
            clip_probs = resnet_probs = vit_probs = empty
            vlm_probs  = None

        pair_scores[(nat_rel, tac_rel)] = {
            opt: {
                "clip":   clip_probs.get(opt),
                "resnet": resnet_probs.get(opt),
                "vit":    vit_probs.get(opt),
                "vlm_ft": vlm_probs.get(opt) if vlm_probs else None,
            }
            for opt in ALL_OPTIONS
        }

    return pair_scores


# ---------------------------------------------------------------------------
# Build rows
# ---------------------------------------------------------------------------
_MODEL_TO_COL = {"clip": "clip_score", "resnet": "resnet_score",
                 "vit": "vit_score", "vlm_ft": "vlm_ft_score"}


def build_rows(records: list[dict], pair_scores: dict, models: list[str],
               flag_by: str = "model") -> list[dict]:
    """
    flag_by='model' : use model consensus vs turker label (original behaviour).
    flag_by='vf'    : flag rows where 0 < vf < 0.5 (partial positive signal,
                      labelled negative) — no model scores included.
    """
    rows = []
    for rec in records:
        vf  = rec.get("vote_fraction")
        opt = rec["option_id"]

        if flag_by == "vf":
            conflict  = _conflict_flag_vf(vf)
            consensus = "N/A"
            default   = str(rec["label"])
            row = {
                "pair_id":         rec["pair_id"],
                "option":          opt,
                "task_family":     rec.get("task_family", ""),
                "turker_vf":       round(vf, 4) if vf is not None else "",
                "current_label":   rec["label"],
                "conflict_flag":   conflict,
                "default_label":   default,
                "user_label":      "",
                "nat_path":        rec["natural_image"],
                "tac_path":        rec["tactile_image"],
            }
        else:
            key        = (rec["natural_image"], rec["tactile_image"])
            all_scores = pair_scores.get(key, {}).get(opt, {
                "clip": None, "resnet": None, "vit": None, "vlm_ft": None})
            scores    = {m: all_scores[m] for m in models}
            turker    = _turker_label(vf)
            consensus = _model_consensus(scores)
            conflict  = _conflict_flag(turker, consensus)
            default   = _default_label(turker, rec["label"], consensus)

            row = {
                "pair_id":        rec["pair_id"],
                "option":         opt,
                "task_family":    rec.get("task_family", ""),
                "turker_vf":      round(vf, 3) if vf is not None else "",
                "current_label":  rec["label"],
            }
            for m in models:
                row[_MODEL_TO_COL[m]] = scores[m]
            row.update({
                "model_consensus": consensus,
                "conflict_flag":   conflict,
                "default_label":   default,
                "user_label":      "",
                "nat_path":        rec["natural_image"],
                "tac_path":        rec["tactile_image"],
            })

        rows.append(row)

    # Sort: conflicts first, then by option, then pair_id
    rows.sort(key=lambda r: (not r["conflict_flag"], r["option"], r["pair_id"]))
    return rows


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------
_CONFLICT_FILL    = PatternFill("solid", fgColor="FFD7D7")   # light red
_BORDERLINE_FILL  = PatternFill("solid", fgColor="FFF3CD")   # light amber
_HEADER_FONT      = Font(bold=True)


def write_excel(rows: list[dict], out_path: Path) -> None:
    df = pd.DataFrame(rows)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Scores")
        ws = writer.sheets["Scores"]

        # Bold headers
        for cell in ws[1]:
            cell.font = _HEADER_FONT

        # Colour rows by conflict / borderline
        col_names = list(df.columns)
        conflict_col_idx = col_names.index("conflict_flag") + 1   # 1-based
        label_col_idx    = col_names.index("user_label") + 1

        for row_idx, row_data in enumerate(rows, start=2):
            fill = None
            if row_data["conflict_flag"]:
                fill = _CONFLICT_FILL
            elif row_data.get("model_consensus") == "SPLIT":
                fill = _BORDERLINE_FILL
            if fill:
                for cell in ws[row_idx]:
                    cell.fill = fill

        # Highlight user_label column header in green
        ws.cell(1, label_col_idx).fill = PatternFill("solid", fgColor="C6EFCE")

        # Column widths
        col_widths = {
            "pair_id": 40, "option": 18, "task_family": 22,
            "turker_vf": 10, "current_label": 13,
            "clip_score": 11, "resnet_score": 13, "vit_score": 10, "vlm_ft_score": 13,
            "model_consensus": 16, "conflict_flag": 13, "default_label": 20,
            "user_label": 12, "nat_path": 50, "tac_path": 50,
        }
        for i, col_name in enumerate(col_names, start=1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = col_widths.get(col_name, 15)

    print(f"  Saved Excel → {out_path}")


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------
_MODEL_DISPLAY = {"clip": "CLIP Probe", "resnet": "ResNet-50",
                  "vit": "ViT-B/16", "vlm_ft": "VLM Fine-tuned"}


def write_summary(rows: list[dict], split: str, out_path: Path, models: list[str]) -> None:
    lines = [
        f"TactileEval Scoring Summary — {split} split",
        f"Models: " + ", ".join(_MODEL_DISPLAY[m] for m in models),
        "",
        f"{'Option':<22} {'Total':>7} {'Conflicts':>10} {'Conflict%':>10} {'Pos':>5} {'Neg':>5}",
        "-" * 60,
    ]
    for opt in ALL_OPTIONS:
        opt_rows = [r for r in rows if r["option"] == opt]
        n_total    = len(opt_rows)
        n_conflict = sum(1 for r in opt_rows if r["conflict_flag"])
        n_pos      = sum(1 for r in opt_rows if r["current_label"] == 1)
        n_neg      = n_total - n_pos
        pct        = f"{100*n_conflict/n_total:.1f}%" if n_total else "—"
        lines.append(f"  {opt:<20} {n_total:>7} {n_conflict:>10} {pct:>10} {n_pos:>5} {n_neg:>5}")

    n_total_all    = len(rows)
    n_conflict_all = sum(1 for r in rows if r["conflict_flag"])
    lines += [
        "-" * 60,
        f"  {'TOTAL':<20} {n_total_all:>7} {n_conflict_all:>10}",
        "",
        "Conflict flag = model consensus disagrees with turker label, OR model consensus is SPLIT.",
        "Review all conflict rows in scored_{split}.xlsx — fill 'user_label' column (0 or 1).",
        "Leave user_label blank for rows you agree with (default_label will be used).",
        "",
        "Next step:",
        f"  python code/review/prepare_inspection.py --split {split}",
        f"  (copies flagged pair images into review/{split}/inspection/ for visual review)",
    ]
    out_path.write_text("\n".join(lines))
    print(f"  Saved summary → {out_path}")
    print("\n".join(lines[4:]))  # print the table to stdout too


def write_summary_vf(rows: list[dict], split: str, out_path: Path) -> None:
    from models.constants import ALL_OPTIONS as _ALL_OPTS
    all_opts = sorted({r["option"] for r in rows})
    lines = [
        f"TactileEval Scoring Summary — {split} split",
        f"Flag method: vf-based (0 < vote_fraction < 0.5 — partial positive signal, labelled negative)",
        "",
        f"{'Option':<22} {'Total':>7} {'Flagged':>10} {'Flagged%':>10} {'Pos':>5} {'Neg':>5}",
        "-" * 60,
    ]
    for opt in all_opts:
        opt_rows   = [r for r in rows if r["option"] == opt]
        n_total    = len(opt_rows)
        n_flagged  = sum(1 for r in opt_rows if r["conflict_flag"])
        n_pos      = sum(1 for r in opt_rows if r["current_label"] == 1)
        n_neg      = n_total - n_pos
        pct        = f"{100*n_flagged/n_total:.1f}%" if n_total else "—"
        lines.append(f"  {opt:<20} {n_total:>7} {n_flagged:>10} {pct:>10} {n_pos:>5} {n_neg:>5}")

    n_total_all   = len(rows)
    n_flagged_all = sum(1 for r in rows if r["conflict_flag"])
    lines += [
        "-" * 60,
        f"  {'TOTAL':<20} {n_total_all:>7} {n_flagged_all:>10}",
        "",
        "Flagged = 0 < vote_fraction < 0.5 (at least one turker saw a defect, majority did not).",
        "These are potential missed positives. Review images and set user_label=1 if defect is present.",
        "Leave user_label blank to keep current label (0).",
        "",
        "Next step:",
        f"  python code/review/prepare_inspection.py --split {split}",
        f"  (copies flagged pair images into review/{split}/inspection/ for visual review)",
    ]
    out_path.write_text("\n".join(lines))
    print(f"  Saved summary → {out_path}")
    print("\n".join(lines[4:]))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--split",    required=True, choices=["test", "val", "train"],
                        help="Dataset split to score")
    parser.add_argument("--models",   nargs="+", default=["clip", "resnet", "vit", "vlm_ft"],
                        choices=["clip", "resnet", "vit", "vlm_ft"],
                        help="Which models to run (default: all four; ignored when --flag-by vf)")
    parser.add_argument("--flag-by",  default="model", choices=["model", "vf"],
                        help="Conflict detection method: 'model' = model consensus (default); "
                             "'vf' = flag rows where 0 < vote_fraction < 0.5 (no model needed)")
    parser.add_argument("--device",   default=None,
                        help="Force device: 'cuda' or 'cpu' (default: auto)")
    parser.add_argument("--out-dir",  default=None,
                        help="Output directory (default: review/{split}/)")
    args = parser.parse_args()

    out_dir = Path(args.out_dir) if args.out_dir else REVIEW_DIR / args.split
    out_dir.mkdir(parents=True, exist_ok=True)

    records = load_split(args.split)
    print(f"Loaded {len(records)} records from {args.split}.jsonl")

    stem = f"scored_{args.split}"

    if args.flag_by == "vf":
        print(f"Flag method: vf-based (0 < vf < 0.5)  |  Split: {args.split}  |  No models run")
        rows = build_rows(records, {}, [], flag_by="vf")
        write_excel(rows, out_dir / f"{stem}.xlsx")
        write_summary_vf(rows, args.split, out_dir / "scoring_summary.txt")
        return

    # --- model-based path ---
    models = args.models
    run_vlm = "vlm_ft" in models

    import torch
    device = args.device or ("cuda" if torch.cuda.is_available() else "cpu")
    if device == "cpu" and run_vlm:
        print("[warn] No GPU detected — dropping VLM Fine-tuned from the model list. "
              "Pass --device cuda to override.")
        models = [m for m in models if m != "vlm_ft"]
        run_vlm = False

    print(f"Device: {device}  |  Split: {args.split}  |  Models: {', '.join(models)}")

    # Load only the selected models
    clip   = CLIPProbeEvaluator(MODELS_DIR / "clip_probe", device=device) if "clip" in models else None
    resnet = TwoStreamEvaluator(MODELS_DIR / "resnet50_twostream" / "resnet50_evaluator.pt", device=device) if "resnet" in models else None
    vit    = TwoStreamEvaluator(MODELS_DIR / "vit_b16_twostream"  / "vit_b_16_evaluator.pt",  device=device) if "vit" in models else None
    vlm    = VLMEvaluator(lora_checkpoint=str(VLM_CKPT), device=device) if run_vlm else None

    pair_scores = score_split(records, clip, resnet, vit, vlm)
    rows = build_rows(records, pair_scores, models, flag_by="model")
    write_excel(rows, out_dir / f"{stem}.xlsx")
    write_summary(rows, args.split, out_dir / "scoring_summary.txt", models)


if __name__ == "__main__":
    main()
